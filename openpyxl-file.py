import openpyxl

sb = openpyxl.load_workbook('Soul Breaks.xlsx')

# Print Sheet Names
print(sb.sheetnames)

ws = sb.active
print(ws['A1'])